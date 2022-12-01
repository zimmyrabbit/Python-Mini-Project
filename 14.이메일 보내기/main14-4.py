from info import info
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb = load_workbook(r'14.이메일 보내기\email_address.xlsx', data_only=True)
load_ws = load_wb.active

for i in range(1,load_ws.max_row + 1) :
    recv_email_value = load_ws.cell(i,1).value
    print(f'recv_email_value : {recv_email_value}')

    try:
        send_email = info.naver_id()
        send_pwd = info.naver_pw()

        recv_email = recv_email_value

        smtp_name = 'smtp.naver.com'
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] = '엑셀에서 메일 주소 읽어 자동 전송'
        msg['From'] = send_email
        msg['To'] = recv_email

        text = '''
            메일 전송 테스트 입니다.
        '''

        msg.attach(MIMEText(text))

        s = smtplib.SMTP(smtp_name, smtp_port)
        s.starttls()
        s.login(send_email, send_pwd)
        s.sendmail(send_email, recv_email, msg.as_string())
        s.quit()
    except:
        print(f'에러 : {recv_email_value}')