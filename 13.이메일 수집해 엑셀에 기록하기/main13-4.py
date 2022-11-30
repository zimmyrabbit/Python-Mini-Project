import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://n.news.naver.com/sports/qatar2022/article/477/0000397625'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'
    , 'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)

result = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

result = list(set(result))

try :
    wb = load_workbook(r'13.이메일 수집해 엑셀에 기록하기\email.xlsx', data_only=False, headers=False)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for i in result:
    sheet.append([i])

wb.save(r'13.이메일 수집해 엑셀에 기록하기\email.xlsx')
